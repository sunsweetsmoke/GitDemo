import openpyxl


class HomePageData:
    #test_homepage_data  =  [
    #    {"firstname":"Ana","email":"cozilovyou@gmail.com","password":"123134","gender":"Female"},
    #    {"firstname":"Dima","email":"metoo@gmail.com","password":"897897","gender":"Male"}
    #    ]

    @staticmethod
    def getTestDataFromXls(test_case_name):
        bookhere = openpyxl.load_workbook("C:\\trainAT\\openpyxlsDemo.xlsx")
        sheet = bookhere.active
        datafroxls = {}

        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i,column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    datafroxls[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [datafroxls]

